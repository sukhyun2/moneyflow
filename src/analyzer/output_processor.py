import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from dateutil.relativedelta import relativedelta

def create_summary_by_month(pdf: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """가공된 데이터에서 타입, 소분류 데이터를 나누는 부분"""
    # 첫 번째 데이터프레임 (대분류가 빈 값인 것들)
    pdf_type = pdf[pdf.index.get_level_values('대분류') == ''].reset_index()

    # 두 번째 데이터프레임 (대분류는 있고 소분류가 빈 값인 것들, 단 '식비'는 소분류까지 포함)
    pdf_small = pdf[
        (
            # 식비
            (pdf.index.get_level_values('대분류') == '식비') &
            (pdf.index.get_level_values('소분류') != '') &
            (pdf.index.get_level_values('내용') == '')
        ) | (
            # others
            (pdf.index.get_level_values('대분류') != '') &
            (pdf.index.get_level_values('대분류') != '식비') &
            (pdf.index.get_level_values('소분류') == '')
        )
    ].reset_index()
    pdf_small = pdf_small.sort_values(by=['타입', '대분류', '소분류', '금액합계'])

    return pdf_type, pdf_small

def split_and_join_summary_by_month(pdf: pd.DataFrame) -> pd.DataFrame:
    """데이터를 월별로 나누고 다시 조인하는 부분 (전월 대비 계산을 위해)"""
    pdf_tmp = pdf[['month', '타입', '대분류', '소분류', '내용', '금액합계']]
    # pdf_tmp을 month 단위로 쪼개서 full outer join하기
    # 1. month별로 데이터 분리
    months = pdf_tmp['month'].unique()

    # 2. 각 month별로 데이터프레임 생성
    month_dfs = {}
    for month in months:
        df_month = pdf_tmp[pdf_tmp['month'] == month].copy()
        # month 컬럼 제거 (join할 때 불필요)
        df_month = df_month.drop('month', axis=1)
        # 컬럼명에 month 정보 추가
        df_month = df_month.rename(columns={
            '금액합계': f'금액합계_{month}'
        })
        month_dfs[month] = df_month

    # 3. Full outer join 수행
    # 첫 번째 month를 기준으로 시작
    result_df = None
    join_keys = ['타입', '대분류', '소분류', '내용']

    for i, (month, df) in enumerate(month_dfs.items()):
        if i == 0:
            result_df = df
        else:
            result_df = pd.merge(result_df, df, on=join_keys, how='outer')

    # 4. 결과 정리 (NaN 값을 0으로 채우고 정렬)
    result_df = result_df.fillna(0)
    result_df = result_df.sort_values(['타입', '대분류', '소분류'])

    return result_df


def filter_target_month_summary(pdf_summ_type, pdf_summ_small, config):
    """타겟월과 전월만 필터링하는 함수"""
    # month
    target_month_ym = config['target_month'].strftime('%Y-%m')
    previous_month_ym = (config['target_month'] - relativedelta(months=1)).strftime('%Y-%m')

    pdf_summ_type_tar_prev = split_and_join_summary_by_month(
        pdf_summ_type[pdf_summ_type['month'].isin([target_month_ym, previous_month_ym])]
    )

    pdf_summ_small_tar_prev = split_and_join_summary_by_month(
        pdf_summ_small[pdf_summ_small['month'].isin([target_month_ym, previous_month_ym])]
    )

    return pdf_summ_type_tar_prev, pdf_summ_small_tar_prev

def create_dataframes_with_separators(dataframes: list[pd.DataFrame]) -> pd.DataFrame:
    """DataFrame 리스트를 받아서 각 DataFrame 사이에 여백을 주어 결합하는 함수"""
    if not dataframes:
        return pd.DataFrame()

    # 빈 DataFrame들 제거
    non_empty_dfs = [df for df in dataframes if len(df) > 0]

    if not non_empty_dfs:
        return pd.DataFrame()

    if len(non_empty_dfs) == 1:
        return non_empty_dfs[0]

    # 결합할 DataFrame 리스트 준비
    combined_dfs = []

    for i, df in enumerate(non_empty_dfs):
        # DataFrame 추가
        combined_dfs.append(df)

        # 마지막 DataFrame이 아니면 여백 추가
        if i < len(non_empty_dfs) - 1:
            # 현재 DataFrame의 컬럼 구조에 맞춰 빈 행 생성
            empty_row = pd.DataFrame([[None] * len(df.columns)], columns=df.columns)
            combined_dfs.append(empty_row)

    # 모든 DataFrame을 결합
    result_df = pd.concat(combined_dfs, ignore_index=True)
    return result_df

def add_dataframe_to_excel(
    df: pd.DataFrame,
    file_path: str,
    sheet_name: str,
    overwrite_sheet: bool = True,
    include_index: bool = False,
    create_file_if_not_exists: bool = True
) -> bool:
    """
    DataFrame을 지정된 엑셀 파일의 특정 시트에 저장하는 범용 함수

    Args:
        df: 저장할 pandas DataFrame
        file_path: 엑셀 파일 전체 경로
        sheet_name: 저장할 시트 이름
        overwrite_sheet: 기존 시트가 있을 때 덮어쓸지 여부 (기본값: True)
        include_index: DataFrame의 인덱스를 포함할지 여부 (기본값: False)
        create_file_if_not_exists: 파일이 없을 때 새로 생성할지 여부 (기본값: True)

    Returns:
        bool: 저장 성공 여부
    """
    try:
        # 빈 DataFrame 체크
        if df.empty:
            print(f"Warning: Empty DataFrame provided for sheet '{sheet_name}'")
            return False

        # 기존 시트 삭제 (overwrite_sheet가 True인 경우)
        if overwrite_sheet:
            try:
                wb = load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                    print(f"Existing sheet '{sheet_name}' deleted")
                wb.save(file_path)
            except FileNotFoundError:
                if not create_file_if_not_exists:
                    print(f"File not found: {file_path}")
                    return False
                print(f"File not found. Will create new file: {file_path}")

        # DataFrame을 엑셀에 저장
        with pd.ExcelWriter(
            file_path,
            engine='openpyxl',
            mode='a' if overwrite_sheet else 'w',
            if_sheet_exists='new' if overwrite_sheet else 'error'
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=include_index)

        print(f"DataFrame saved successfully to '{file_path}' sheet '{sheet_name}'")
        print(f"Data shape: {df.shape}")
        return True

    except FileExistsError:
        print(f"Error: Sheet '{sheet_name}' already exists and overwrite_sheet is False")
        return False
    except Exception as e:
        print(f"Error saving DataFrame to Excel: {e}")
        return False


def process_asset_data(config) -> pd.DataFrame:
    """asset.xlsx 파일을 읽어와서 pivot table로 변환하고 엑셀 파일에 저장하는 함수"""
    # asset.xlsx 파일 읽기
    asset_file_path = config['input_path'] + '/' + config['asset_file_name']

    try:
        pdf = pd.read_excel(asset_file_path)
    except FileNotFoundError:
        print(f"Asset file not found: {asset_file_path}")
        return pd.DataFrame()

    # 필수 컬럼 확인
    required_columns = ['날짜', '카테고리', '세부항목', '금액']
    missing_columns = [col for col in required_columns if col not in pdf.columns]
    if missing_columns:
        print(f"Missing required columns in asset.xlsx: {missing_columns}")
        return pd.DataFrame()

    # 날짜 컬럼을 datetime으로 변환 후 yyyy-mm 형식으로 변경
    pdf['날짜'] = pd.to_datetime(pdf['날짜']).dt.strftime('%Y-%m')

    # pivot table 생성
    pdf_pivot = pdf.pivot_table(
        index=['카테고리', '세부항목'],
        columns='날짜',
        values='금액',
        aggfunc='sum'
    )

    # 출력 파일 경로 설정
    output_file_path = config['output_path'] + '/' + config['output_file_name']
    asset_sheet = 'asset_summary'

    # 기존 asset_summary 시트가 있으면 삭제
    try:
        wb = load_workbook(output_file_path)
        if asset_sheet in wb.sheetnames:
            del wb[asset_sheet]
        wb.save(output_file_path)
    except FileNotFoundError:
        # 파일이 없으면 새로 생성될 것이므로 통과
        pass

    # 엑셀 파일에 새 시트로 저장
    with pd.ExcelWriter(
        output_file_path,
        engine='openpyxl',
        mode='a',
        if_sheet_exists='new'
    ) as writer:
        pdf_pivot.to_excel(writer, sheet_name=asset_sheet)

    print(f"Asset data processed and saved to sheet: {asset_sheet}")
    return pdf_pivot

def auto_adjust_column_width(file_path: str):
    """엑셀 파일의 모든 컬럼 너비를 자동으로 조정하는 함수"""
    try:
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # 각 컬럼의 너비를 자동 조정
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value is not None:
                        # 실제 표시되는 형태의 길이를 계산
                        if isinstance(cell.value, (int, float)):
                            # 숫자인 경우 포맷팅을 고려한 길이 계산
                            if cell.number_format and cell.number_format != 'General':
                                # 포맷이 적용된 경우 대략적인 길이 추정
                                if ',' in cell.number_format:  # 천 단위 구분자가 있는 경우
                                    formatted_value = f"{cell.value:,}"
                                else:
                                    formatted_value = str(cell.value)
                            else:
                                formatted_value = str(cell.value)
                            cell_length = len(formatted_value)
                        else:
                            # 텍스트인 경우
                            cell_length = len(str(cell.value))

                        max_length = max(max_length, cell_length)

                # 너비 조정 (최소 12, 최대 80, 여백 +5)
                adjusted_width = min(max(max_length + 7, 20), 100)
                sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)
        print(f"Column width adjusted for: {file_path}")
        return True

    except Exception as e:
        print(f"Error adjusting column width: {e}")
        return False


def auto_adjust_column_width_to_output(config):
    """config에서 지정된 출력 파일에 컬럼 너비 자동 조정을 적용하는 함수"""
    output_file_path = config['output_path'] + '/' + config['output_file_name']
    return auto_adjust_column_width(output_file_path)


def apply_accounting_format(file_path: str):
    """지정된 엑셀 파일의 모든 시트에서 숫자값을 회계 형식으로 변경하는 함수 (음수는 절댓값으로 변환)"""
    try:
        # 엑셀 파일 로드
        wb = load_workbook(file_path)

        # 회계 형식 정의 (천 단위 구분자, 음수도 양수 형태로 표시)
        accounting_format = "#,##0"

        # 모든 시트 순회
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Processing accounting format for sheet: {sheet_name}")

            # 모든 셀 순회하여 숫자 포맷 적용 및 음수를 절댓값으로 변환
            for row in sheet.iter_rows():
                for cell in row:
                    # 셀에 값이 있고 숫자인 경우에만 처리
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        # 음수인 경우 절댓값으로 변환
                        if cell.value < 0:
                            cell.value = abs(cell.value)
                        # 회계 형식 적용
                        cell.number_format = accounting_format

        # 파일 저장
        wb.save(file_path)
        print(f"Accounting format applied successfully to: {file_path}")
        return True

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return False
    except Exception as e:
        print(f"Error applying accounting format: {e}")
        return False


def apply_accounting_format_to_output(config):
    """config에서 지정된 출력 파일에 회계 형식을 적용하는 함수"""
    output_file_path = config['output_path'] + '/' + config['output_file_name']
    return apply_accounting_format(output_file_path)

def set_font_size_for_all_sheets(file_path: str, font_size: int) -> bool:
    """
    지정된 엑셀 파일의 모든 시트에서 글자 크기를 변경하는 함수

    Args:
        file_path: 엑셀 파일 전체 경로
        font_size: 설정할 글자 크기 (포인트)

    Returns:
        bool: 설정 성공 여부
    """
    try:
        # 엑셀 파일 로드
        wb = load_workbook(file_path)

        # 모든 시트 순회
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Setting font size to {font_size} for sheet: {sheet_name}")

            # 모든 셀 순회하여 폰트 크기 적용
            for row in sheet.iter_rows():
                for cell in row:
                    # 셀에 값이 있는 경우에만 폰트 적용
                    if cell.value is not None:
                        # 기존 폰트의 다른 속성을 유지하면서 크기만 변경
                        current_font = cell.font
                        try:
                            # 기존 폰트 속성들을 안전하게 가져오기
                            font_name = getattr(current_font, 'name', 'Calibri')
                            font_bold = getattr(current_font, 'bold', False)
                            font_italic = getattr(current_font, 'italic', False)
                            font_underline = getattr(current_font, 'underline', 'none')
                            font_color = getattr(current_font, 'color', None)

                            cell.font = Font(
                                name=font_name,
                                size=font_size,
                                bold=font_bold,
                                italic=font_italic,
                                underline=font_underline,
                                color=font_color
                            )
                        except Exception:
                            # 속성 접근에 실패하면 기본 폰트로 크기만 설정
                            cell.font = Font(size=font_size)

        # 파일 저장
        wb.save(file_path)
        print(f"Font size set to {font_size} successfully for: {file_path}")
        return True

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return False
    except Exception as e:
        print(f"Error setting font size: {e}")
        return False

def set_font_size_for_output(config, font_size: int):
    """config에서 지정된 출력 파일에 글자 크기를 적용하는 함수"""
    output_file_path = config['output_path'] + '/' + config['output_file_name']
    return set_font_size_for_all_sheets(output_file_path, font_size)